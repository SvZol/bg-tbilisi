from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, BackgroundTasks
from sqlalchemy.orm import Session
from database import get_db
from models.content import Post, Page, EventResult, EventPhoto
from models.event import Event
from core.security import get_current_admin
from pydantic import BaseModel
from uuid import UUID, uuid4
from typing import Optional
from datetime import datetime
import shutil
import os

router = APIRouter(prefix="/admin", tags=["admin"])

UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# --- Мероприятия ---

class EventCreate(BaseModel):
    title: str
    description: str | None = None
    city: str | None = None
    starts_at: datetime
    ends_at: datetime
    reg_deadline: datetime
    min_team_size: int = 1
    max_team_size: int = 10

class EventOut(BaseModel):
    id: UUID
    title: str
    description: str | None
    city: str | None = None
    starts_at: datetime
    ends_at: datetime
    reg_deadline: datetime
    min_team_size: int
    max_team_size: int
    status: str

    class Config:
        from_attributes = True

class EventEdit(BaseModel):
    title: str | None = None
    description: str | None = None
    city: str | None = None
    min_team_size: int | None = None
    max_team_size: int | None = None

@router.get("/events", response_model=list[EventOut])
def list_all_events(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return db.query(Event).order_by(Event.starts_at.desc()).all()

@router.post("/events", response_model=EventOut)
def create_event(data: EventCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    event = Event(**data.model_dump())
    db.add(event)
    db.commit()
    db.refresh(event)
    return event

@router.patch("/events/{event_id}", response_model=EventOut)
def edit_event(event_id: UUID, data: EventEdit, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Мероприятие не найдено")
    if data.title is not None: event.title = data.title
    if data.description is not None: event.description = data.description
    if data.city is not None: event.city = data.city
    if data.min_team_size is not None: event.min_team_size = data.min_team_size
    if data.max_team_size is not None: event.max_team_size = data.max_team_size
    db.commit()
    db.refresh(event)
    return event

@router.patch("/events/{event_id}/status")
def update_event_status(event_id: UUID, status: str, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Мероприятие не найдено")
    event.status = status
    db.commit()
    return {"ok": True}


class EventReschedule(BaseModel):
    starts_at: datetime
    ends_at: datetime
    reg_deadline: datetime


@router.patch("/events/{event_id}/reschedule")
def reschedule_event(
    event_id: UUID,
    data: EventReschedule,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin)
):
    from models.team import Team
    from models.user import User
    from core.email import send_reschedule_email

    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(404, "Мероприятие не найдено")

    event.starts_at = data.starts_at
    event.ends_at = data.ends_at
    event.reg_deadline = data.reg_deadline
    db.commit()
    db.refresh(event)

    # Собираем email всех участников команд этого мероприятия
    teams = db.query(Team).filter(Team.event_id == event_id).all()
    notified = set()
    for team in teams:
        for member in team.members:
            email = None
            if member.user_id:
                u = db.query(User).filter(User.id == member.user_id).first()
                if u:
                    email = u.email
            elif member.guest_email:
                email = member.guest_email
            if email and email not in notified:
                notified.add(email)
                background_tasks.add_task(
                    _send_safe, send_reschedule_email,
                    email, event.title, team.name,
                    event.starts_at, event.reg_deadline
                )

    return {"ok": True, "notified": len(notified)}


@router.patch("/events/{event_id}/reg-deadline")
def update_reg_deadline(
    event_id: UUID,
    data: dict,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin)
):
    """Update registration deadline silently (no email notifications)."""
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(404, "Мероприятие не найдено")
    from datetime import datetime as dt
    reg_deadline = data.get("reg_deadline")
    if not reg_deadline:
        raise HTTPException(400, "reg_deadline обязателен")
    event.reg_deadline = dt.fromisoformat(reg_deadline)
    db.commit()
    return {"ok": True, "reg_deadline": event.reg_deadline.isoformat()}


@router.post("/events/{event_id}/notify-new")
def notify_new_event(
    event_id: UUID,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin)
):
    from models.user import User
    from core.email import send_new_event_email

    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(404, "Мероприятие не найдено")

    users = db.query(User).filter(User.is_verified == True).all()
    for u in users:
        background_tasks.add_task(
            _send_safe, send_new_event_email,
            u.email, event.title, event.starts_at, event.reg_deadline, str(event_id)
        )

    return {"ok": True, "notified": len(users)}


@router.post("/events/{event_id}/upload-results-pdf")
async def upload_results_pdf(
    event_id: UUID,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    import shutil, uuid as _uuid
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(404, "Мероприятие не найдено")
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "Разрешены только PDF-файлы")

    os.makedirs("uploads/pdfs", exist_ok=True)
    filename = f"{_uuid.uuid4()}.pdf"
    path = f"uploads/pdfs/{filename}"

    # Удаляем старый файл если был
    if event.results_pdf:
        old = f"uploads/pdfs/{event.results_pdf}"
        if os.path.exists(old):
            os.remove(old)

    content = await file.read()
    with open(path, "wb") as f:
        f.write(content)

    event.results_pdf = filename
    db.commit()
    return {"ok": True, "filename": filename}


@router.delete("/events/{event_id}/results-pdf")
def delete_results_pdf(event_id: UUID, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(404, "Мероприятие не найдено")
    if event.results_pdf:
        path = f"uploads/pdfs/{event.results_pdf}"
        if os.path.exists(path):
            os.remove(path)
        event.results_pdf = None
        db.commit()
    return {"ok": True}


def _send_safe(fn, *args):
    try:
        fn(*args)
    except Exception as e:
        print(f"[EMAIL ERROR] {type(e).__name__}: {e}")

# --- Новости ---

class PostCreate(BaseModel):
    title: str
    content: str
    is_published: bool = False

@router.post("/posts")
def create_post(data: PostCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    post = Post(**data.model_dump(), author_id=admin.id)
    db.add(post)
    db.commit()
    db.refresh(post)
    return post

@router.get("/posts/public")
def list_posts_public(db: Session = Depends(get_db)):
    return db.query(Post).filter(Post.is_published == True).order_by(Post.created_at.desc()).all()

@router.get("/posts")
def list_posts(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    return db.query(Post).order_by(Post.created_at.desc()).all()

@router.patch("/posts/{post_id}")
def update_post(post_id: UUID, data: PostCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    post = db.query(Post).filter(Post.id == post_id).first()
    if not post:
        raise HTTPException(status_code=404, detail="Не найдено")
    for key, value in data.model_dump().items():
        setattr(post, key, value)
    post.updated_at = datetime.utcnow()
    db.commit()
    return post

@router.post("/posts/{post_id}/image")
def upload_post_image(
    post_id: UUID,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin)
):
    post = db.query(Post).filter(Post.id == post_id).first()
    if not post:
        raise HTTPException(status_code=404, detail="Не найдено")
    filename = f"{uuid4()}_{file.filename}"
    filepath = os.path.join(UPLOAD_DIR, filename)
    with open(filepath, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    post.image_filename = filename
    post.updated_at = datetime.utcnow()
    db.commit()
    return {"image_filename": filename}

@router.delete("/posts/{post_id}")
def delete_post(post_id: UUID, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    post = db.query(Post).filter(Post.id == post_id).first()
    if not post:
        raise HTTPException(404, "Не найдено")
    if post.image_filename:
        path = os.path.join(UPLOAD_DIR, post.image_filename)
        if os.path.exists(path):
            os.remove(path)
    db.delete(post)
    db.commit()
    return {"ok": True}

# --- Страницы ---

class PageCreate(BaseModel):
    slug: str
    title: str
    content: str
    is_published: bool = False

@router.post("/pages")
def create_page(data: PageCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    page = Page(**data.model_dump())
    db.add(page)
    db.commit()
    db.refresh(page)
    return page

@router.patch("/pages/{page_id}")
def update_page(page_id: UUID, data: PageCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    page = db.query(Page).filter(Page.id == page_id).first()
    if not page:
        raise HTTPException(status_code=404, detail="Не найдено")
    for key, value in data.model_dump().items():
        setattr(page, key, value)
    page.updated_at = datetime.utcnow()
    db.commit()
    return page

# --- Результаты ---

class ResultCreate(BaseModel):
    event_id: UUID
    team_id: UUID
    rank: Optional[str] = None
    score: Optional[str] = None
    notes: Optional[str] = None

@router.post("/results")
def create_result(data: ResultCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    result = EventResult(**data.model_dump())
    db.add(result)
    db.commit()
    db.refresh(result)
    return result

@router.get("/results/{event_id}")
def get_results(event_id: UUID, db: Session = Depends(get_db)):
    return db.query(EventResult).filter(EventResult.event_id == event_id).all()

# --- Фото ---

@router.post("/photos/{event_id}")
def upload_photo(
    event_id: UUID,
    caption: Optional[str] = None,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin)
):
    filename = f"{uuid4()}_{file.filename}"
    filepath = os.path.join(UPLOAD_DIR, filename)
    with open(filepath, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    photo = EventPhoto(event_id=event_id, filename=filename, caption=caption)
    db.add(photo)
    db.commit()
    return {"filename": filename}

@router.get("/photos/{event_id}")
def get_photos(event_id: UUID, db: Session = Depends(get_db)):
    return db.query(EventPhoto).filter(EventPhoto.event_id == event_id).all()


# --- Команды ---

@router.get("/teams/{event_id}")
def get_event_teams(event_id: UUID, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    from models.team import Team, TeamMember
    from models.user import User

    teams = db.query(Team).filter(Team.event_id == event_id).all()

    result = []
    for team in teams:
        members_data = []
        for member in team.members:
            member_dict = {
                "id": str(member.id),
                "user_id": str(member.user_id) if member.user_id else None,
                "guest_name": member.guest_name,
                "guest_email": member.guest_email,
                "role": member.role,
                "is_registered": member.is_registered,
                "display_name": None,
                "display_email": None,
            }
            if member.user_id:
                user = db.query(User).filter(User.id == member.user_id).first()
                if user:
                    member_dict["display_name"] = user.full_name
                    member_dict["display_email"] = user.email
                    member_dict["last_login_at"] = user.last_login_at.isoformat() if user.last_login_at else None
                    member_dict["is_imported"] = bool(user.is_imported)
            else:
                member_dict["display_name"] = member.guest_name
                member_dict["display_email"] = member.guest_email
                member_dict["last_login_at"] = None
                member_dict["is_imported"] = False

            members_data.append(member_dict)

        result.append({
            "id": str(team.id),
            "name": team.name,
            "status": team.status,
            "category": team.category or "adult",
            "members": members_data
        })

    return result

@router.patch("/teams/{team_id}/category")
def set_team_category(team_id: UUID, category: str, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    from models.team import Team
    team = db.query(Team).filter(Team.id == team_id).first()
    if not team:
        raise HTTPException(404, "Команда не найдена")
    if category not in ("adult", "child"):
        raise HTTPException(400, "Допустимые значения: adult, child")
    if category == "child":
        from models.team import TeamMember
        member_count = db.query(TeamMember).filter(TeamMember.team_id == team_id).count()
        if member_count < 2:
            raise HTTPException(400, "Для детского зачёта (Лосята) нужно минимум 2 участника")
    team.category = category
    db.commit()
    return {"id": str(team.id), "name": team.name, "category": team.category}

@router.delete("/teams/{team_id}")
def delete_team(team_id: UUID, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    from models.team import Team, TeamMember
    from models.content import EventResult, TeamQuestionResult
    db.query(TeamQuestionResult).filter(TeamQuestionResult.team_id == team_id).delete()
    db.query(EventResult).filter(EventResult.team_id == team_id).delete()
    db.query(TeamMember).filter(TeamMember.team_id == team_id).delete()
    db.query(Team).filter(Team.id == team_id).delete()
    db.commit()
    return {"ok": True}

@router.delete("/members/{member_id}")
def delete_member(member_id: UUID, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    from models.team import TeamMember
    member = db.query(TeamMember).filter(TeamMember.id == member_id).first()
    if not member:
        raise HTTPException(status_code=404, detail="Не найдено")
    team_id = member.team_id
    db.delete(member)
    db.flush()
    # Если остался 1 участник — автоматически переключаем на взрослый зачёт
    from models.team import Team as _Team
    team = db.query(_Team).filter(_Team.id == team_id).first()
    if team:
        remaining = db.query(TeamMember).filter(TeamMember.team_id == team_id).count()
        if remaining < 2 and team.category == "child":
            team.category = "adult"
    db.commit()
    return {"ok": True}


# --- Вопросы и ответы ---

class QuestionCreate(BaseModel):
    number: int
    text: str
    correct_answer: Optional[str] = None
    max_points: int = 1

class QuestionUpdate(BaseModel):
    text: Optional[str] = None
    correct_answer: Optional[str] = None
    max_points: Optional[int] = None
    is_published: Optional[bool] = None

def _q_out(q) -> dict:
    return {
        "id": str(q.id), "event_id": str(q.event_id), "number": q.number,
        "text": q.text, "correct_answer": q.correct_answer,
        "max_points": q.max_points, "is_published": q.is_published,
        "image_filename": q.image_filename,
    }

class QuestionResultUpsert(BaseModel):
    team_id: UUID
    points_earned: int

@router.post("/events/{event_id}/questions")
def create_question(event_id: UUID, data: QuestionCreate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    from models.content import EventQuestion
    q = EventQuestion(event_id=event_id, **data.model_dump())
    db.add(q)
    db.commit()
    db.refresh(q)
    return _q_out(q)

@router.get("/events/{event_id}/questions")
def get_questions_admin(event_id: UUID, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    from models.content import EventQuestion
    qs = db.query(EventQuestion).filter(EventQuestion.event_id == event_id).order_by(EventQuestion.number).all()
    return [_q_out(q) for q in qs]

@router.patch("/questions/{question_id}")
def update_question(question_id: UUID, data: QuestionUpdate, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    from models.content import EventQuestion
    q = db.query(EventQuestion).filter(EventQuestion.id == question_id).first()
    if not q:
        raise HTTPException(status_code=404, detail="Не найдено")
    for key, value in data.model_dump(exclude_none=True).items():
        setattr(q, key, value)
    db.commit()
    return _q_out(q)

@router.post("/questions/{question_id}/image")
async def upload_question_image(
    question_id: UUID, file: UploadFile = File(...),
    db: Session = Depends(get_db), admin=Depends(get_current_admin)
):
    from models.content import EventQuestion
    import uuid as uuid_mod
    q = db.query(EventQuestion).filter(EventQuestion.id == question_id).first()
    if not q:
        raise HTTPException(404, "Не найдено")
    ext = os.path.splitext(file.filename or "")[1].lower() or ".jpg"
    filename = f"q_{uuid_mod.uuid4().hex}{ext}"
    path = os.path.join("uploads", filename)
    with open(path, "wb") as f:
        content = await file.read()
        f.write(content)
    if q.image_filename:
        try: os.remove(os.path.join("uploads", q.image_filename))
        except: pass
    q.image_filename = filename
    db.commit()
    return {"image_filename": filename}

@router.delete("/events/{event_id}/questions")
def delete_all_questions(event_id: UUID, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    from models.content import EventQuestion, TeamQuestionResult
    q_ids = [q.id for q in db.query(EventQuestion).filter(EventQuestion.event_id == event_id).all()]
    if q_ids:
        db.query(TeamQuestionResult).filter(TeamQuestionResult.question_id.in_(q_ids)).delete(synchronize_session=False)
    deleted = db.query(EventQuestion).filter(EventQuestion.event_id == event_id).delete()
    db.commit()
    return {"deleted": deleted}


@router.delete("/questions/{question_id}")
def delete_question(question_id: UUID, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    from models.content import EventQuestion, TeamQuestionResult
    db.query(TeamQuestionResult).filter(TeamQuestionResult.question_id == question_id).delete()
    db.query(EventQuestion).filter(EventQuestion.id == question_id).delete()
    db.commit()
    return {"ok": True}

@router.post("/questions/{question_id}/results")
def upsert_question_result(question_id: UUID, data: QuestionResultUpsert, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    from models.content import TeamQuestionResult
    existing = db.query(TeamQuestionResult).filter(
        TeamQuestionResult.question_id == question_id,
        TeamQuestionResult.team_id == data.team_id
    ).first()
    if existing:
        existing.points_earned = data.points_earned
    else:
        existing = TeamQuestionResult(question_id=question_id, team_id=data.team_id, points_earned=data.points_earned)
        db.add(existing)
    db.commit()
    return {"ok": True}

@router.get("/events/{event_id}/question-results")
def get_question_results_admin(event_id: UUID, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    from models.content import EventQuestion, TeamQuestionResult
    from models.team import Team
    qs = db.query(EventQuestion).filter(EventQuestion.event_id == event_id).order_by(EventQuestion.number).all()
    teams = db.query(Team).filter(Team.event_id == event_id).all()
    results = db.query(TeamQuestionResult).join(EventQuestion).filter(EventQuestion.event_id == event_id).all()
    result_map = {f"{r.question_id}|{r.team_id}": r.points_earned for r in results}
    return {
        "questions": [_q_out(q) for q in qs],
        "teams": [{"id": str(t.id), "name": t.name, "category": t.category or "adult"} for t in teams],
        "results": result_map,
    }

# --- Шаблон Excel для результатов ---

@router.get("/events/{event_id}/results-template")
async def generate_results_template(
    event_id: UUID,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin)
):
    import openpyxl, io
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse
    from models.content import EventQuestion
    from models.team import Team

    questions = db.query(EventQuestion).filter(EventQuestion.event_id == event_id).order_by(EventQuestion.number).all()
    teams = db.query(Team).filter(Team.event_id == event_id).all()

    if not questions:
        raise HTTPException(400, "Сначала импортируйте вопросы (КП)")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "результаты"

    from openpyxl.utils import get_column_letter

    dark = PatternFill("solid", start_color="292524")
    gray = PatternFill("solid", start_color="D6D3D1")
    orange = PatternFill("solid", start_color="FED7AA")
    purple = PatternFill("solid", start_color="EDE9FE")
    orange_hdr = PatternFill("solid", start_color="EA580C")
    purple_hdr = PatternFill("solid", start_color="7C3AED")

    # Строка 1 — служебная: первые 3 ячейки пустые, затем UUID команды на каждую пару колонок
    # Строка 2 — видимая шапка: КП, Тип, Эталон, затем «Название / ответ», «Название / балл» для каждой команды
    row1 = ['__id__', '', '']
    row2 = ['КП', 'Тип', 'Правильный ответ']
    for t in teams:
        row1 += [str(t.id), '']   # UUID, пустая (для колонки балла)
        row2 += [t.name + ' — ответ', t.name + ' — балл']
    ws.append(row1)
    ws.append(row2)

    # Форматирование строки 1 (служебная — серая, мелкий шрифт)
    for cell in ws[1]:
        cell.font = Font(size=8, color="78716C")
        cell.fill = gray

    # Форматирование строки 2 (шапка)
    for col_idx, cell in enumerate(ws[2], 1):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = dark
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Ширина колонок
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 22
    for i in range(len(teams)):
        ans_col = get_column_letter(4 + i * 2)
        pts_col = get_column_letter(5 + i * 2)
        ws.column_dimensions[ans_col].width = 20
        ws.column_dimensions[pts_col].width = 8

    # Группируем вопросы по КП
    kp_map: dict = {}
    for q in questions:
        if q.number < 100:
            kp_map.setdefault(q.number, {})['zadanie'] = q
        else:
            kp_map.setdefault(q.number - 100, {})['zadacha'] = q

    for kp_num in sorted(kp_map.keys()):
        kp = kp_map[kp_num]
        label = f"КП-{kp_num:02d}"

        if 'zadanie' in kp:
            q = kp['zadanie']
            row = [label, 'задание КП', q.correct_answer or ''] + ['' for _ in range(len(teams) * 2)]
            ws.append(row)
            r = ws.max_row
            # Первые 3 ячейки — оранжевые
            for c in range(1, 4):
                ws.cell(r, c).fill = orange_hdr
                ws.cell(r, c).font = Font(bold=True, color="FFFFFF")
            # Ячейки команд — светло-оранжевые
            for c in range(4, 4 + len(teams) * 2):
                ws.cell(r, c).fill = orange

        if 'zadacha' in kp:
            q = kp['zadacha']
            row = ['', 'задача', q.correct_answer or ''] + ['' for _ in range(len(teams) * 2)]
            ws.append(row)
            r = ws.max_row
            for c in range(1, 4):
                ws.cell(r, c).fill = purple_hdr
                ws.cell(r, c).font = Font(bold=True, color="FFFFFF")
            for c in range(4, 4 + len(teams) * 2):
                ws.cell(r, c).fill = purple

    # Высота строк шапки
    ws.row_dimensions[1].height = 15
    ws.row_dimensions[2].height = 35
    ws.freeze_panes = 'D3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=results_template.xlsx"}
    )


# --- Импорт из Excel ---

@router.post("/events/{event_id}/import-kp-excel")
async def import_kp_excel(
    event_id: UUID,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin)
):
    import openpyxl, io
    from models.content import EventQuestion, TeamQuestionResult

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "Не удалось открыть файл Excel")

    ws = None
    for name in wb.sheetnames:
        if 'КП' in name or 'кп' in name.lower():
            ws = wb[name]
            break
    if ws is None:
        raise HTTPException(400, "Лист 'Все КП' не найден")

    rows = list(ws.iter_rows(values_only=True))

    # Удаляем старые результаты и вопросы
    old_q_ids = [q.id for q in db.query(EventQuestion).filter(EventQuestion.event_id == event_id).all()]
    if old_q_ids:
        db.query(TeamQuestionResult).filter(TeamQuestionResult.question_id.in_(old_q_ids)).delete(synchronize_session=False)
    db.query(EventQuestion).filter(EventQuestion.event_id == event_id).delete()

    SKIP_TYPES = {'тип кп', 'header', ''}
    created = 0
    for row in rows[1:]:  # skip header
        if not row[4]:
            continue
        kp_str = str(row[4]).strip()
        if not kp_str.upper().startswith('КП'):
            continue
        kp_type = str(row[1]).strip() if row[1] else ""
        if kp_type.lower() in SKIP_TYPES:
            continue
        try:
            num = int(kp_str.split('-')[1])
        except Exception:
            continue

        task_text = str(row[8]).strip() if row[8] else None
        answer = str(row[9]).strip() if row[9] else None
        if answer == 'None':
            answer = None

        # Задание КП (если есть текст задания или ответ)
        if task_text and task_text != 'None':
            db.add(EventQuestion(
                event_id=event_id, number=num, kp_type=kp_type,
                text=task_text, correct_answer=answer, max_points=1
            ))
            created += 1
        elif answer and answer != 'None':
            db.add(EventQuestion(
                event_id=event_id, number=num, kp_type=kp_type,
                text=f"КП-{num:02d}", correct_answer=answer, max_points=1
            ))
            created += 1
        elif kp_type in ('Старт', 'Финиш'):
            # Старт и Финиш без задания — всё равно создаём
            db.add(EventQuestion(
                event_id=event_id, number=num, kp_type=kp_type,
                text=kp_type, correct_answer=None, max_points=0
            ))
            created += 1

        # Задача (если есть)
        puzzle_text = str(row[11]).strip() if row[11] else None
        puzzle_answer = str(row[12]).strip() if row[12] else None
        if puzzle_text and puzzle_text != 'None':
            if puzzle_answer == 'None':
                puzzle_answer = None
            db.add(EventQuestion(
                event_id=event_id, number=num + 100, kp_type='Задача',
                text=f"Задача: {puzzle_text}",
                correct_answer=puzzle_answer, max_points=1
            ))
            created += 1

    db.commit()
    return {"created": created, "message": f"Создано {created} вопросов (задания + задачи)"}


# --- Импорт команд из Google-формы (Excel) ---

@router.post("/events/{event_id}/import-teams-excel")
async def import_teams_excel(
    event_id: UUID,
    file: UploadFile = File(...),
    background_tasks: BackgroundTasks = None,
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    from models.team import Team, TeamMember
    from models.user import User
    from core.security import hash_password
    from core.email import send_invite_email
    import openpyxl, io, secrets, string

    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(404, "Мероприятие не найдено")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    created = []
    skipped = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # пустая строка
            continue

        team_name = str(row[1]).strip() if row[1] else None
        member_count = int(row[2]) if row[2] and str(row[2]).isdigit() else None
        category_raw = str(row[3]).strip() if row[3] else ""
        contact = str(row[4]).strip() if row[4] else None

        if not team_name:
            continue

        # Определяем зачёт
        category = "child" if "дети" in category_raw.lower() else "adult"

        # Проверяем дубликат
        existing = db.query(Team).filter(Team.event_id == event_id, Team.name == team_name).first()
        if existing:
            skipped.append(f"{team_name} (уже есть)")
            continue

        # Определяем email (содержит @ и точку после @, не Telegram-ссылка)
        import re as _re
        email = None
        telegram = None
        if contact and _re.match(r'^[^@]+@[^@]+\.[^@]+$', contact) and "t.me" not in contact:
            email = contact
        else:
            telegram = contact

        # Создаём или находим пользователя по email
        user = None
        temp_password = None
        if email:
            user = db.query(User).filter(User.email == email).first()
            if not user:
                # Генерируем временный пароль
                alphabet = string.ascii_letters + string.digits
                temp_password = "".join(secrets.choice(alphabet) for _ in range(10))
                user = User(
                    email=email,
                    password_hash=hash_password(temp_password),
                    full_name=team_name,  # временное имя
                    is_verified=True,
                    role="user",
                    is_imported=True,
                )
                db.add(user)
                db.flush()

        # Генерируем уникальный код приглашения
        import secrets as _sec, string as _str
        _alph = _str.ascii_uppercase + _str.digits
        while True:
            invite_code = "".join(_sec.choice(_alph) for _ in range(10))
            if not db.query(Team).filter(Team.invite_code == invite_code).first():
                break

        # Создаём команду
        team = Team(
            event_id=event_id,
            created_by=user.id if user else admin.id,
            name=team_name,
            category=category,
            member_count=member_count,
            captain_phone=telegram,  # telegram в поле телефона если нет email
            description=f"Импорт из формы. Контакт: {contact}" if contact else "Импорт из формы",
            invite_code=invite_code,
        )
        db.add(team)
        db.flush()

        # Добавляем капитана
        captain = TeamMember(
            team_id=team.id,
            user_id=user.id if user else None,
            guest_name=team_name if not user else None,
            role="captain",
            is_registered=user is not None,
        )
        db.add(captain)

        # Добавляем анонимных участников согласно member_count
        if member_count and member_count > 1:
            for idx in range(2, member_count + 1):
                db.add(TeamMember(
                    team_id=team.id,
                    guest_name=f"Участник {idx}",
                    role="member",
                    is_registered=False,
                ))

        # Отправляем письмо если есть email и новый пользователь
        if email and temp_password and background_tasks:
            background_tasks.add_task(
                _send_safe, send_invite_email,
                email, team_name, event.title, temp_password, str(event_id)
            )

        # Генерируем сообщение для Telegram если нет email
        site_url = os.getenv("FRONTEND_URL", "https://tbi-ssector.run")
        cat_label = "Лосята (детский зачёт)" if category == "child" else "Лоси (взрослый зачёт)"
        claim_url = f"{site_url}/join?code={invite_code}"

        tg_message = None
        if telegram and not email:
            tg_message = (
                f"Привет! 👋 Ваша команда «{team_name}» ({cat_label}) зарегистрирована на игру «{event.title}».\n\n"
                f"Чтобы стать владельцем команды на сайте, перейдите по ссылке:\n"
                f"{claim_url}\n\n"
                f"Ссылка одноразовая — после привязки команды она перестанет работать."
            )

        created.append({
            "team": team_name,
            "category": category,
            "members": member_count,
            "contact": contact,
            "email_sent": bool(email and temp_password),
            "telegram": telegram,
            "telegram_message": tg_message,
            "invite_code": invite_code,
            "claim_url": claim_url,
        })

    db.commit()
    return {
        "created": len(created),
        "skipped": len(skipped),
        "teams": created,
        "skipped_names": skipped,
    }


@router.post("/events/{event_id}/import-results-excel")
async def import_results_excel(
    event_id: UUID,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin)
):
    import openpyxl, io
    from models.content import EventQuestion, TeamQuestionResult
    from models.team import Team

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "Не удалось открыть файл Excel")

    # Определяем формат:
    # - Шаблон (наш): лист 'результаты', строка 1 начинается с '__id__'
    # - Оригинальный файл: лист 'проверка'
    use_id_template = False
    if 'проверка' in wb.sheetnames:
        ws = wb['проверка']
    elif 'результаты' in wb.sheetnames:
        ws = wb['результаты']
        # Проверяем что это наш шаблон (первая ячейка = '__id__')
        first_cell = ws.cell(1, 1).value
        if str(first_cell).strip() == '__id__':
            use_id_template = True
        else:
            raise HTTPException(400, "Лист 'результаты' найден, но формат не распознан. Загрузите файл с листом 'проверка' или шаблон скачанный из системы.")
    else:
        raise HTTPException(400, "Не найден лист 'проверка'. Убедитесь что загружаете правильный файл результатов.")

    rows = list(ws.iter_rows(values_only=True))

    # Загружаем команды из БД
    db_teams = db.query(Team).filter(Team.event_id == event_id).all()
    team_map_by_name = {t.name.lower().strip(): t for t in db_teams}
    team_map_by_id = {str(t.id): t for t in db_teams}

    team_col_start = 3  # cols 3+ = team data
    # team_entries: list of (Team|None, ans_col_idx, pts_col_idx)
    team_entries: list[tuple] = []
    team_names: list[str] = []

    if use_id_template:
        # Строка 0 (row[0]): col[0]='__id__', col[3], col[5], ... = UUID команды
        # Строка 1 (row[1]): col[0]='КП', col[3]='Имя — ответ', col[4]='Имя — балл', ...
        # Данные с строки 2
        id_row = rows[0] if len(rows) > 0 else []
        rows = rows[2:]  # пропускаем обе строки шапки
        col = team_col_start
        while col < len(id_row):
            uid = str(id_row[col]).strip() if id_row[col] else ''
            if not uid or uid in ('None', 'nan', ''):
                break
            t = team_map_by_id.get(uid)
            team_entries.append((t, col, col + 1))  # ответ в col, балл в col+1
            team_names.append(t.name if t else uid)
            col += 2
    else:
        # Оригинальный формат: ищем строку с именами
        team_names: list[str] = []
        for row in rows:
            if row[2] and 'талон' in str(row[2]) and row[3] and str(row[3]).strip() not in ('nan', 'None', ''):
                team_names = [str(row[i]).strip() for i in range(team_col_start, len(row)) if row[i] and str(row[i]).strip() not in ('nan', 'None', '')]
                break
        if not team_names:
            raise HTTPException(400, "Не найдена строка с названиями команд")
        # Для оригинального формата: одна колонка на команду (только балл, без ответа)
        team_entries = [(team_map_by_name.get(n.lower().strip()), team_col_start + i, None) for i, n in enumerate(team_names)]

    # Авто-создаём недостающие команды (только для оригинального формата)
    auto_created = 0
    if not use_id_template:
        if 'регистрация' in wb.sheetnames:
            reg_ws = wb['регистрация']
            for reg_row in reg_ws.iter_rows(values_only=True):
                if not reg_row[2]:
                    continue
                name = str(reg_row[2]).strip()
                if not name or name in ('None', 'nan'):
                    continue
                if name.lower() not in team_map_by_name:
                    new_team = Team(event_id=event_id, name=name, status='registered', created_by=admin.id)
                    db.add(new_team)
                    db.flush()
                    team_map_by_name[name.lower()] = new_team
                    auto_created += 1
        else:
            for name in team_names:
                if name.lower() not in team_map_by_name:
                    new_team = Team(event_id=event_id, name=name, status='registered', created_by=admin.id)
                    db.add(new_team)
                    db.flush()
                    team_map_by_name[name.lower()] = new_team
                    auto_created += 1
        # Обновляем team_entries после авто-создания
        team_entries = [(team_map_by_name.get(n.lower().strip()), team_col_start + i, None) for i, n in enumerate(team_names)]

    # Читаем лист 'ответы' — реальные ответы команд и правильные ответы
    # answers_map[(kp_num, q_type)] = {team_name_lower: answer_str}
    # correct_answers_map[(kp_num, q_type)] = answer_str  (col[2] = эталон)
    answers_map: dict[tuple, dict] = {}
    correct_answers_map: dict[tuple, str] = {}

    if 'ответы' in wb.sheetnames:
        ans_ws = wb['ответы']
        ans_rows = list(ans_ws.iter_rows(values_only=True))

        # Найти строку с именами команд (та же логика)
        ans_team_names: list[str] = []
        for r in ans_rows:
            if r[2] and 'талон' in str(r[2]) and r[3] and str(r[3]).strip() not in ('nan', 'None', ''):
                ans_team_names = [str(r[i]).strip() for i in range(team_col_start, len(r)) if r[i] and str(r[i]).strip() not in ('nan', 'None', '')]
                break

        cur_kp: Optional[int] = None
        for r in ans_rows:
            c0 = r[0]
            c1 = str(r[1]).strip() if r[1] else ""
            if c0 and str(c0).strip().startswith('КП'):
                try:
                    cur_kp = int(str(c0).strip().split('-')[1])
                except Exception:
                    cur_kp = None
            if cur_kp is None:
                continue
            if 'задание' in c1.lower():
                q_type = 'zadanie'
                q_number = cur_kp
            elif 'задача' in c1.lower():
                q_type = 'zadacha'
                q_number = cur_kp + 100
            else:
                continue

            # Правильный ответ из col[2]
            etalon = r[2]
            if etalon is not None and str(etalon).strip() not in ('None', 'nan', ''):
                correct_answers_map[(cur_kp, q_type)] = str(etalon).strip()

            # Ответы команд
            team_answers: dict = {}
            for i, tname in enumerate(ans_team_names):
                col_idx = team_col_start + i
                val = r[col_idx] if col_idx < len(r) else None
                ans_str = str(val).strip() if val is not None and str(val).strip() not in ('None', 'nan', '') else None
                team_answers[tname.lower()] = ans_str
            answers_map[(cur_kp, q_type)] = team_answers

    # Загружаем вопросы из БД
    questions = db.query(EventQuestion).filter(EventQuestion.event_id == event_id).all()
    q_map = {q.number: q for q in questions}

    if not q_map:
        raise HTTPException(400, "Сначала импортируйте вопросы (КП)")

    # Обновляем correct_answer на вопросах из листа ответы (точнее эталона)
    for (kp_num, q_type), ans_str in correct_answers_map.items():
        q_number = kp_num if q_type == 'zadanie' else kp_num + 100
        q = q_map.get(q_number)
        if q:
            q.correct_answer = ans_str

    # Удаляем старые результаты
    q_ids = [q.id for q in questions]
    db.query(TeamQuestionResult).filter(TeamQuestionResult.question_id.in_(q_ids)).delete(synchronize_session=False)

    unmatched: list[str] = []
    created = 0
    current_kp: Optional[int] = None

    for row in rows:
        col0 = row[0]
        col1 = str(row[1]).strip() if row[1] else ""

        if col0 and str(col0).strip().startswith('КП'):
            try:
                current_kp = int(str(col0).strip().split('-')[1])
            except Exception:
                current_kp = None

        if current_kp is None:
            continue

        if 'задание' in col1.lower():
            q_number = current_kp
            q_type = 'zadanie'
        elif 'задача' in col1.lower():
            q_number = current_kp + 100
            q_type = 'zadacha'
        else:
            continue

        question = q_map.get(q_number)
        if not question:
            continue

        team_answers_for_row = answers_map.get((current_kp, q_type), {})

        for i, (db_team, ans_col, pts_col) in enumerate(team_entries):
            # ans_col — колонка с ответом, pts_col — с баллом (None если оригинальный формат)
            pts_idx = pts_col if pts_col is not None else ans_col
            ans_idx = ans_col

            if pts_idx >= len(row):
                continue
            score_val = row[pts_idx]
            try:
                score = int(score_val) if score_val is not None else 0
            except Exception:
                score = 0

            if not db_team:
                tname = team_names[i] if i < len(team_names) else f"col{i}"
                if tname not in unmatched:
                    unmatched.append(tname)
                continue

            # Ответ: из шаблона — своя колонка, из оригинала — из листа ответы
            if pts_col is not None:
                # Шаблон: ответ в ans_col
                ans_val = row[ans_idx] if ans_idx < len(row) else None
                team_ans = str(ans_val).strip() if ans_val is not None and str(ans_val).strip() not in ('None', 'nan', '') else None
            else:
                team_ans = team_answers_for_row.get(db_team.name.lower().strip())

            db.add(TeamQuestionResult(
                question_id=question.id,
                team_id=db_team.id,
                points_earned=score,
                team_answer=team_ans,
            ))
            created += 1

    # Собираем имена ДО commit (после commit объекты становятся expired)
    teams_found = [str(e[0].name) for e in team_entries if e[0] is not None]

    db.commit()
    return {
        "created": int(created),
        "auto_created_teams": int(auto_created),
        "unmatched_teams": list(unmatched),
        "teams_found": teams_found,
    }


# --- Публичный доступ к вопросам ---

@router.get("/public/events/{event_id}/questions")
def get_questions_public(event_id: UUID, db: Session = Depends(get_db)):
    from models.content import EventQuestion, TeamQuestionResult
    from models.team import Team
    qs = db.query(EventQuestion).filter(
        EventQuestion.event_id == event_id,
        EventQuestion.is_published == True
    ).order_by(EventQuestion.number).all()
    if not qs:
        return {"questions": [], "teams": [], "results": {}}
    teams = db.query(Team).filter(Team.event_id == event_id).all()
    results = db.query(TeamQuestionResult).join(EventQuestion).filter(EventQuestion.event_id == event_id).all()
    result_map = {f"{r.question_id}|{r.team_id}": r.points_earned for r in results}
    return {
        "questions": [_q_out(q) for q in qs],
        "teams": [{"id": str(t.id), "name": t.name, "category": t.category or "adult"} for t in teams],
        "results": result_map,
    }

# --- Страницы (правила, информация) ---

class PageUpsert(BaseModel):
    title: str
    content: str
    is_published: bool = True

@router.get("/public/pages/{slug}")
def get_page_public(slug: str, db: Session = Depends(get_db)):
    page = db.query(Page).filter(Page.slug == slug, Page.is_published == True).first()
    if not page:
        raise HTTPException(404, "Страница не найдена")
    return {"slug": page.slug, "title": page.title, "content": page.content}

@router.get("/pages/{slug}")
def get_page_admin(slug: str, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    page = db.query(Page).filter(Page.slug == slug).first()
    if not page:
        raise HTTPException(404, "Страница не найдена")
    return {"slug": page.slug, "title": page.title, "content": page.content, "is_published": page.is_published}

@router.post("/pages/{slug}/image")
async def upload_page_image(
    slug: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin=Depends(get_current_admin),
):
    import uuid as _uuid
    allowed = {'.jpg', '.jpeg', '.png', '.gif', '.webp'}
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed:
        raise HTTPException(400, "Разрешены только изображения")
    os.makedirs("uploads/pages", exist_ok=True)
    filename = f"{_uuid.uuid4()}{ext}"
    path = f"uploads/pages/{filename}"
    content = await file.read()
    with open(path, "wb") as f:
        f.write(content)
    return {"filename": filename, "url": f"/uploads/pages/{filename}"}

@router.put("/pages/{slug}")
def upsert_page(slug: str, data: PageUpsert, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    from datetime import timezone
    page = db.query(Page).filter(Page.slug == slug).first()
    if page:
        page.title = data.title
        page.content = data.content
        page.is_published = data.is_published
        page.updated_at = datetime.now(timezone.utc)
    else:
        page = Page(slug=slug, title=data.title, content=data.content, is_published=data.is_published)
        db.add(page)
    db.commit()
    return {"slug": page.slug, "title": page.title, "content": page.content, "is_published": page.is_published}


# --- Скорборд ---

@router.get("/events/{event_id}/scoreboard")
def get_scoreboard(event_id: UUID, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    from models.content import TeamQuestionResult, EventQuestion
    from models.team import Team

    teams = db.query(Team).filter(Team.event_id == event_id).all()
    results = (
        db.query(TeamQuestionResult)
        .join(EventQuestion)
        .filter(EventQuestion.event_id == event_id)
        .all()
    )

    score_map: dict = {}
    for r in results:
        key = str(r.team_id)
        score_map[key] = score_map.get(key, 0) + (r.points_earned or 0)

    rows = []
    for team in teams:
        rows.append({
            "id": str(team.id),
            "name": team.name,
            "category": team.category or "adult",
            "score": score_map.get(str(team.id), 0),
        })

    adult = sorted([r for r in rows if r["category"] != "child"], key=lambda x: -x["score"])
    child = sorted([r for r in rows if r["category"] == "child"], key=lambda x: -x["score"])

    for i, r in enumerate(adult): r["rank"] = i + 1
    for i, r in enumerate(child): r["rank"] = i + 1

    return {"adult": adult, "child": child}


# --- Публикация результатов ---

@router.post("/events/{event_id}/publish-results")
def publish_results(event_id: UUID, background_tasks: BackgroundTasks, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    from models.content import EventQuestion, EventResult
    from models.team import Team
    from models.user import User
    from core.email import send_results_email

    updated = db.query(EventQuestion).filter(EventQuestion.event_id == event_id).update({"is_published": True})
    db.commit()

    event = db.query(Event).filter(Event.id == event_id).first()
    if event:
        from models.content import TeamQuestionResult, EventQuestion as EQ
        teams = db.query(Team).filter(Team.event_id == event_id).all()
        all_pts = db.query(TeamQuestionResult).join(EQ).filter(EQ.event_id == event_id).all()

        # Считаем баллы для каждой команды
        score_map: dict = {}
        for p in all_pts:
            key = str(p.team_id)
            score_map[key] = score_map.get(key, 0) + (p.points_earned or 0)

        # Вычисляем место внутри каждого зачёта
        for cat in ("adult", "child"):
            cat_teams = [t for t in teams if (t.category or "adult") == cat]
            cat_teams.sort(key=lambda t: -score_map.get(str(t.id), 0))
            for i, t in enumerate(cat_teams):
                score_map[f"rank_{t.id}"] = i + 1

        notified = set()
        for team in teams:
            score = score_map.get(str(team.id), 0)
            rank = score_map.get(f"rank_{team.id}")
            for member in team.members:
                email = None
                if member.user_id:
                    u = db.query(User).filter(User.id == member.user_id).first()
                    if u:
                        email = u.email
                elif member.guest_email:
                    email = member.guest_email
                if email and email not in notified:
                    notified.add(email)
                    background_tasks.add_task(
                        _send_safe, send_results_email,
                        email, event.title, team.name, rank, score, str(event_id)
                    )

    return {"published": updated}

@router.post("/events/{event_id}/unpublish-results")
def unpublish_results(event_id: UUID, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    from models.content import EventQuestion
    updated = db.query(EventQuestion).filter(EventQuestion.event_id == event_id).update({"is_published": False})
    db.commit()
    return {"unpublished": updated}


# --- Управление пользователями ---

@router.get("/users")
def list_users(db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    from models.user import User
    users = db.query(User).order_by(User.created_at.desc()).all()
    return [
        {
            "id": str(u.id),
            "email": u.email,
            "full_name": u.full_name,
            "role": u.role,
            "is_verified": u.is_verified,
            "created_at": u.created_at.isoformat() if u.created_at else None,
            "last_login_at": u.last_login_at.isoformat() if u.last_login_at else None,
            "is_imported": bool(u.is_imported),
        }
        for u in users
    ]

@router.post("/teams/{team_id}/invite-code")
def generate_team_invite_code(team_id: UUID, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    """Generate (or regenerate) an invite code for a team so admin can share the claim link."""
    team = db.query(Team).filter(Team.id == team_id).first()
    if not team:
        raise HTTPException(404, "Команда не найдена")
    import secrets as _sec, string as _str
    _alph = _str.ascii_uppercase + _str.digits
    while True:
        code = "".join(_sec.choice(_alph) for _ in range(10))
        if not db.query(Team).filter(Team.invite_code == code).first():
            break
    team.invite_code = code
    db.commit()
    return {"invite_code": code}


@router.patch("/users/{user_id}/role")
def set_user_role(user_id: UUID, data: dict, db: Session = Depends(get_db), admin=Depends(get_current_admin)):
    from models.user import User
    role = data.get("role")
    if role not in ("user", "admin"):
        raise HTTPException(400, "Роль должна быть 'user' или 'admin'")
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(404, "Пользователь не найден")
    if str(user.id) == str(admin.id):
        raise HTTPException(400, "Нельзя изменить свою роль")
    user.role = role
    db.commit()
    return {"id": str(user.id), "email": user.email, "role": user.role}
